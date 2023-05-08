import openpyxl as xl


wb = xl.load_workbook('transactions.xlsx')  #loads the excel sheet
sheet = wb['Sheet1'] #this is name of sheet in excel
cell = sheet['a1'] # co-ordinate of cell

for row in range(2, sheet.max_row + 1): # row 2 till last row
    cell = sheet.cell(row, 3) # stores values from row 3 into 'cell'
    corrected_price = (cell.value * 0.9) #multiplies row 3 values by 0.9
    corrected_price_cell = sheet.cell(row, 4) #puts corrct value on new row
    corrected_price_cell.value = corrected_price


    wb.save('transactionsNew.xlsx')